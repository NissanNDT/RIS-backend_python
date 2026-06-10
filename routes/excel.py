import os
from datetime import datetime, date
from fastapi import APIRouter, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config.db import db_pool

router = APIRouter()

@router.get("/excel/{id_incident}")
def generate_excel(id_incident: int):
    # Fetch incident
    with db_pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM incident WHERE id = %s", [id_incident])
            incident = cur.fetchone()
            if not incident:
                raise HTTPException(status_code=404, detail="Incident not found")
            
            # Fetch incident format
            cur.execute("SELECT * FROM incident_format WHERE id_incident = %s", [id_incident])
            format_data = cur.fetchone()
            
            format_id = format_data["id"] if format_data else None
            
            # Fetch related details if format exists
            factors = []
            countermeasures = []
            participants = []
            hazard_backgrounds = []
            hazard_background = None
            intervening_factors = []
            
            if format_id:
                cur.execute("SELECT * FROM factor_tree WHERE id_incident_format = %s ORDER BY id ASC", [format_id])
                factors = cur.fetchall()
                
                cur.execute("SELECT * FROM countermeasure_plan WHERE id_incident_format = %s", [format_id])
                countermeasures = cur.fetchall()
                
                cur.execute("SELECT * FROM analysis_participant WHERE id_incident_format = %s", [format_id])
                participants = cur.fetchall()
                
                # Obtener todos los campos del hazard_background (antecedentes de peligro)
                cur.execute("SELECT * FROM hazard_background WHERE id_incident_format = %s ORDER BY id ASC", [format_id])
                hazard_backgrounds = cur.fetchall()
                hazard_background = hazard_backgrounds[0] if hazard_backgrounds else None
                
                cur.execute("SELECT * FROM intervening_factors WHERE id_incident_format = %s", [format_id])
                intervening_factors = cur.fetchall()

            # Get names for plant, area, supervisors
            cur.execute("SELECT name FROM area WHERE id = %s", [incident["id_area"]])
            area = cur.fetchone()
            
            cur.execute("SELECT name FROM plant WHERE id = %s", [incident["id_plant"]])
            plant = cur.fetchone()
            
            cur.execute("SELECT full_name FROM users WHERE id = %s", [incident["id_responsible_user"]])
            resp_user = cur.fetchone()
            
            cur.execute("SELECT full_name FROM users WHERE id = %s", [incident["id_general_sv"]])
            gen_sv = cur.fetchone()
            
            cur.execute("SELECT full_name FROM users WHERE id = %s", [incident["id_junior"]])
            junior = cur.fetchone()
            
            cur.execute("SELECT * FROM cost_center")
            cost_centers = cur.fetchall()
            
            cur.execute("SELECT * FROM control_hierarchy")
            hierarchies = cur.fetchall()
            
            cur.execute("SELECT * FROM verification_method")
            methods = cur.fetchall()

    # Load template
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Anexo 1 Reporte de Incidente.xlsx")
    if not os.path.exists(template_path):
        # Fallback to current working directory
        template_path = "Anexo 1 Reporte de Incidente.xlsx"
        
    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Template Excel file not found or corrupted: {str(e)}")
        
    ws = wb["REPORTE"] if "REPORTE" in wb.sheetnames else wb.active

    def set_cell(coord, val):
        from openpyxl.cell.cell import Cell
        cell = ws[coord]
        if isinstance(cell, Cell):
            cell.value = val
        else:
            for r in ws.merged_cells.ranges:
                if coord in r:
                    ws.cell(row=r.min_row, column=r.min_col, value=val)
                    return
            cell.value = val

    # Fill Header
    set_cell("B2", incident["incident_folio"] or "")
    
    # Date formatting
    inc_date = incident["date"]
    if isinstance(inc_date, (datetime, date)):
        set_cell("B3", inc_date.strftime("%Y-%m-%d"))
        set_cell("I17", inc_date.day)
        set_cell("L17", inc_date.month)
        set_cell("O17", inc_date.year)
    else:
        set_cell("B3", str(inc_date) if inc_date else "")

    set_cell("B4", area["name"] if area else "")
    if plant:
        set_cell("F4", plant["name"])
        set_cell("V15", plant["name"])

    # Fills
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    level_str = incident["level"] or ""
    if "G" in level_str: ws["AB4"].fill = red_fill
    if "U" in level_str: ws["AD4"].fill = red_fill
    if "R" in level_str: ws["AF4"].fill = red_fill
    if "FR1" in level_str: ws["AH4"].fill = red_fill
    if "FR0" in level_str: ws["AJ4"].fill = red_fill

    set_cell("G7", incident["incident_mechanism"] or "")
    set_cell("G9", incident["injury"] or "")

    if format_data:
        set_cell("I13", format_data["employee_name"] or "")
        set_cell("AC13", format_data["employee_age"] or "")
        set_cell("AL13", format_data["employee_payroll_number"] or "")
        set_cell("I19", format_data["employee_position"] or "")
        set_cell("X19", format_data["employee_distribution"] or "")
        set_cell("AL17", format_data["employee_seniority"] or "")
        set_cell("AX17", format_data["employee_seniority_in_position"] or "")
        
        if format_data["employee_type"] == "Sindicalizado":
            ws["AT19"].fill = red_fill
        elif format_data["employee_type"] == "No sindicalizado":
            ws["BB19"].fill = red_fill
            
        set_cell("AV13", format_data["accident_shift"] or "")
        set_cell("I23", format_data["sv_seniority"] or "")
        set_cell("AC23", format_data["sv_seniority_in_position"] or "")
        set_cell("AW23", format_data["number_of_staff_under_sv"] or "")
        set_cell("K25", format_data["attending_doctor"] or "")
        set_cell("AG25", format_data["recovery_forecast"] or "")

    if incident["id_cost_center"]:
        cc = next((c for c in cost_centers if c["id"] == incident["id_cost_center"]), None)
        set_cell("I15", cc["name"] if cc else "")

    if incident["time"]:
        set_cell("AB17", incident["time"])

    if resp_user: set_cell("I21", resp_user["full_name"])
    if gen_sv: set_cell("Z21", gen_sv["full_name"])
    if junior: set_cell("AS21", junior["full_name"])

    set_cell("C28", incident["description"] or "")

    # =========================================================
    # CAUSA RAÍZ (row 55 fija en template)
    # =========================================================
    set_cell("C55", incident["root_cause"] or "")

    # =========================================================
    # ANTECEDENTES DE PELIGRO - Sección 7
    # Referencia visual: Reporte de incidente IMEX.xls sección 7
    # Se escribe ANTES de insertar filas del árbol de factores.
    # openpyxl desplaza automáticamente los valores cuando se llama insert_rows.
    # =========================================================
    def _col_num(col_str):
        """Convierte letras de columna a número (A=1, Z=26, AA=27, etc.)"""
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col_str)

    if hazard_background:
        # Row 61: ¿Se han presentado incidentes FR1 previos? + Revisión horizontal
        try:
            if hazard_background.get("previous_fr1_incidents_presented") is True:
                ws.cell(row=61, column=_col_num("R")).fill = black_fill  # Checkbox SI (Col 18)
            elif hazard_background.get("previous_fr1_incidents_presented") is False:
                ws.cell(row=61, column=_col_num("H")).fill = black_fill  # Checkbox NO (Col 8)
        except Exception:
            pass
        try:
            if hazard_background.get("horizontal_review") is True:
                ws.cell(row=61, column=_col_num("AL")).fill = black_fill # Checkbox SI (Col 38)
                ws.cell(row=62, column=_col_num("AT")).value = hazard_background.get("horizontal_review_comment") or ""
            elif hazard_background.get("horizontal_review") is False:
                ws.cell(row=61, column=_col_num("AQ")).fill = black_fill # Checkbox NO (Col 43)
        except Exception:
            pass

        # Row 66: ¿Existen procesos/áreas con potencial de incidente?
        try:
            if hazard_background.get("existing_processes_or_areas_potential_for_incident") is True:
                ws.cell(row=66, column=_col_num("R")).fill = black_fill  # Checkbox SI (Col 18)
                ws.cell(row=66, column=_col_num("AJ")).value = hazard_background.get("processes_or_areas_potential_for_incident") or ""
            elif hazard_background.get("existing_processes_or_areas_potential_for_incident") is False:
                ws.cell(row=66, column=_col_num("H")).fill = black_fill  # Checkbox NO (Col 8)
        except Exception:
            pass

        # Row 71: Riesgo identificado / nueva evaluación necesaria
        try:
            if hazard_background.get("risk_assessed_and_identified") is True:
                ws.cell(row=71, column=_col_num("R")).fill = black_fill  # Checkbox SI (Col 18)
            elif hazard_background.get("risk_assessed_and_identified") is False:
                ws.cell(row=71, column=_col_num("H")).fill = black_fill  # Checkbox NO (Col 8)
        except Exception:
            pass
        try:
            if hazard_background.get("new_risk_assessment_needed") is True:
                ws.cell(row=71, column=_col_num("AW")).fill = black_fill # Checkbox SI (Col 49)
            elif hazard_background.get("new_risk_assessment_needed") is False:
                ws.cell(row=71, column=_col_num("AM")).fill = black_fill # Checkbox NO (Col 39)
        except Exception:
            pass

        # Row 76: Taller de limitaciones funcionales (default NO) + Safety Dojo
        try:
            ws.cell(row=76, column=_col_num("R")).fill = black_fill  # Checkbox NO (Col 18)
        except Exception:
            pass
        safety_dojo_date = hazard_background.get("safety_dojo_reception_date")
        try:
            if safety_dojo_date:
                if isinstance(safety_dojo_date, (datetime, date)):
                    ws.cell(row=76, column=_col_num("AB")).value = safety_dojo_date.strftime("%d/%m/%Y")
                else:
                    ws.cell(row=76, column=_col_num("AB")).value = str(safety_dojo_date)
            else:
                ws.cell(row=76, column=_col_num("BA")).fill = black_fill # Checkbox NO (Col 53)
        except Exception:
            pass

        # Row 77: Fecha Genba Dojo
        genba_dojo_date = hazard_background.get("genba_dojo_reception_date")
        try:
            if genba_dojo_date:
                if isinstance(genba_dojo_date, (datetime, date)):
                    ws.cell(row=77, column=_col_num("AB")).value = genba_dojo_date.strftime("%d/%m/%Y")
                else:
                    ws.cell(row=77, column=_col_num("AB")).value = str(genba_dojo_date)
            else:
                ws.cell(row=77, column=_col_num("BA")).fill = black_fill # Checkbox NO (Col 53)
        except Exception:
            pass

        # Row 81: Tipo de negligencia + Reporte laboral
        try:
            negligence = hazard_background.get("negligence_type")
            if negligence == "Negligencia consciente":
                ws.cell(row=81, column=_col_num("Z")).fill = black_fill  # Checkbox Consciente (Col 26)
            elif negligence == "Negligencia no consciente":
                ws.cell(row=81, column=_col_num("AH")).fill = black_fill # Checkbox No Consciente (Col 34)
        except Exception:
            pass
        try:
            if hazard_background.get("labor_report") is True:
                ws.cell(row=81, column=_col_num("BA")).fill = black_fill # Checkbox SI (Col 53)
            elif hazard_background.get("labor_report") is False:
                ws.cell(row=81, column=_col_num("AR")).fill = black_fill # Checkbox NO (Col 44)
        except Exception:
            pass

    # Intervening factors checks (row 81) - Acto / Condición insegura
    has_acto = any(f.get("name") and "acto" in f["name"].lower() for f in intervening_factors)
    has_condicion = any(f.get("name") and ("condicion" in f["name"].lower() or "condición" in f["name"].lower()) for f in intervening_factors)
    try:
        if has_acto:
            ws.cell(row=81, column=_col_num("H")).fill = black_fill  # Checkbox Acto (Col 8)
        if has_condicion:
            ws.cell(row=81, column=_col_num("P")).fill = black_fill  # Checkbox Condición (Col 16)
    except Exception:
        pass

    # Intervening factors list (filas 49-53 fijas del template)
    # Estas filas están ANTES del árbol de factores (row 36-43) así que no se desplazan
    row_idx = 49
    for f in intervening_factors:
        if row_idx <= 53:
            set_cell(f"D{row_idx}", f.get("name") or "")
            row_idx += 1

    # =========================================================
    # ÁRBOL DE FACTORES - Generación dinámica con inserción de filas
    # Referencia visual: Reporte de incidente IMEX.xls sección 4
    # Columnas: Lesión | División (4M) | Factor | Subtipo | Descripción
    # Se inserta DESPUÉS de escribir los antecedentes para que openpyxl
    # desplace automáticamente las filas ya escritas.
    # =========================================================
    # Group factors by category (4M)
    categories = [
        {"name": "Mano de Obra", "match": ["mano de obra", "mano"], "factors": []},
        {"name": "Método", "match": ["metodo", "método"], "factors": []},
        {"name": "Maquinaria", "match": ["maquinaria", "máquinaria"], "factors": []},
        {"name": "Materiales", "match": ["material", "materiales"], "factors": []}
    ]

    for f in factors:
        f_4m = str(f.get("4m") or f.get("m4") or f.get("m4_name") or f.get("m4_name_es") or "").lower().strip()
        matched = False
        for cat in categories:
            if any(m in f_4m for m in cat["match"]):
                cat["factors"].append(f)
                matched = True
                break
        if not matched:
            categories[0]["factors"].append(f) # fallback

    for cat in categories:
        if not cat["factors"]:
            cat["factors"].append({
                "factor": "",
                "control_point": "",
                "standard": "",
                "actual": "",
                "met_standard": None,
                "met_safety": None,
                "comments": ""
            })

    # Escribir y dar estilo a las cabeceras en las filas 38-39 para que se visualice como tabla
    headers_config = [
        {"col_start": 3, "col_end": 7, "text": "DIVISIÓN (4M)"},
        {"col_start": 8, "col_end": 11, "text": "SITUACIÓN ACTUAL"},
        {"col_start": 12, "col_end": 15, "text": "FACTOR"},
        {"col_start": 16, "col_end": 19, "text": "PUNTO DE CONTROL"},
        {"col_start": 20, "col_end": 21, "text": "ESTÁNDAR"},
        {"col_start": 22, "col_end": 23, "text": "CUMPLE NORMA"},
        {"col_start": 24, "col_end": 25, "text": "CUMPLE SEGURIDAD"},
        {"col_start": 26, "col_end": 33, "text": "COMENTARIOS"}
    ]
    
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Azul oscuro corporativo
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(border_style="thin", color="000000")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for h in headers_config:
        # Fusión de filas 38 y 39
        ws.merge_cells(start_row=38, end_row=39, start_column=h["col_start"], end_column=h["col_end"])
        for r in [38, 39]:
            for c in range(h["col_start"], h["col_end"] + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = header_align
                if r == 38 and c == h["col_start"]:
                    cell.value = h["text"]
                    cell.font = header_font

    total_factors_count = sum(len(cat["factors"]) for cat in categories)
    
    # Cada factor ocupa 2 filas. El template tiene capacidad para 8 filas de factor (4 categorías x 2 filas = 8 filas).
    extra_factor_rows = max(0, 2 * total_factors_count - 8)
    if extra_factor_rows > 0:
        factor_insert_at = 48 # En el template original, la fila 48 es "5. FACTORES QUE INTERVIENEN"

        # Paso 1: Recopilar y quitar merges en o debajo del punto de inserción
        factor_merges_to_shift = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= factor_insert_at:
                factor_merges_to_shift.append((
                    mr.min_row, mr.max_row,
                    mr.min_col, mr.max_col
                ))
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row, mr.max_row, mr.min_col, mr.max_col) in factor_merges_to_shift:
                ws.merged_cells.ranges.discard(mr)

        # Paso 2: Insertar filas en blanco
        ws.insert_rows(factor_insert_at, extra_factor_rows)

        # Paso 3: Re-aplicar merges desplazados
        for (min_r, max_r, min_c, max_c) in factor_merges_to_shift:
            ws.merge_cells(
                start_row=min_r + extra_factor_rows,
                end_row=max_r + extra_factor_rows,
                start_column=min_c,
                end_column=max_c
            )

        # Paso 4: Copiar estilos de la última fila de factores (fila 47 en template original)
        from copy import copy
        for idx in range(extra_factor_rows):
            r_idx = factor_insert_at + idx
            ws.row_dimensions[r_idx].height = ws.row_dimensions[47].height
            for col_idx in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=47, column=col_idx)
                dest_cell = ws.cell(row=r_idx, column=col_idx)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.alignment = copy(src_cell.alignment)
                    dest_cell.number_format = src_cell.number_format
                    dest_cell.protection = copy(src_cell.protection)

    # Estilos para celdas de datos
    data_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Rellenar filas del Árbol de Factores
    current_factor_idx = 0
    for cat in categories:
        for j, f in enumerate(cat["factors"]):
            factor_offset = 2 * current_factor_idx
            r_idx = 40 + factor_offset
            
            division = cat["name"] if j == 0 else ""
            factor_val = f.get("factor") or ""
            subtipo = f.get("control_point") or ""
            descripcion = f.get("comments") or ""
            actual_val = f.get("actual") or ""
            standard_val = f.get("standard") or ""
            met_std = "SÍ" if f.get("met_standard") in [True, "SÍ", "SI", "true", 1] else ("NO" if f.get("met_standard") in [False, "NO", "false", 0] else "")
            met_saf = "SÍ" if f.get("met_safety") in [True, "SÍ", "SI", "true", 1] else ("NO" if f.get("met_safety") in [False, "NO", "false", 0] else "")

            # Escribir en celdas
            set_cell(f"C{r_idx}", division)        # División / Categoría 4M
            set_cell(f"H{r_idx}", actual_val)      # Situación actual
            set_cell(f"L{r_idx}", factor_val)      # Factor
            set_cell(f"P{r_idx}", subtipo)         # Subtipo / Punto de control
            set_cell(f"T{r_idx}", standard_val)    # Estándar
            set_cell(f"V{r_idx}", met_std)         # Cumple norma
            set_cell(f"X{r_idx}", met_saf)         # Cumple seguridad
            set_cell(f"Z{r_idx}", descripcion)     # Descripción / Comentarios

            # Aplicar bordes, alineación y fuente para toda la fila de datos
            for r in [r_idx, r_idx+1]:
                for c in range(3, 34):
                    cell = ws.cell(row=r, column=c)
                    cell.border = data_border
                    cell.alignment = Alignment(horizontal="center" if c in [3, 22, 23, 24, 25] else "left", vertical="center", wrap_text=True)
                    cell.font = Font(name="Arial", size=9)

            # Combinar celdas de forma vertical (2 filas por factor)
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=3, end_column=7)   # C:G
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=8, end_column=11)  # H:K
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=12, end_column=15) # L:O
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=16, end_column=19) # P:S
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=20, end_column=21) # T:U
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=22, end_column=23) # V:W
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=24, end_column=25) # X:Y
            ws.merge_cells(start_row=r_idx, end_row=r_idx+1, start_column=26, end_column=33) # Z:AG
            
            current_factor_idx += 1

    # Calcular desplazamiento total por filas insertadas en árbol de factores
    factor_shift = extra_factor_rows

    # Countermeasure Plans (starting at Row 87, desplazado por factor_shift)
    template_capacity = 7
    start_row = 87 + factor_shift
    extra_rows = len(countermeasures) - template_capacity if len(countermeasures) > template_capacity else 0

    if extra_rows > 0:
        insert_at = start_row + template_capacity

        merges_to_shift = []
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= insert_at:
                merges_to_shift.append((
                    mr.min_row, mr.max_row,
                    mr.min_col, mr.max_col
                ))

        # Remove all merge ranges that need to be shifted
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row, mr.max_row, mr.min_col, mr.max_col) in merges_to_shift:
                ws.merged_cells.ranges.discard(mr)

        # --- Step 2: Insert the blank rows ---
        ws.insert_rows(insert_at, extra_rows)

        # --- Step 3: Re-apply shifted merges ---
        for (min_r, max_r, min_c, max_c) in merges_to_shift:
            ws.merge_cells(
                start_row=min_r + extra_rows, end_row=max_r + extra_rows,
                start_column=min_c, end_column=max_c
            )

        # --- Step 4: Apply styles and merges to the newly inserted rows ---
        from copy import copy
        for idx in range(extra_rows):
            r_idx = insert_at + idx

            # Copy row height
            ws.row_dimensions[r_idx].height = ws.row_dimensions[start_row].height

            # Copy cell styles/borders from the first template countermeasure row
            for col_idx in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=start_row, column=col_idx)
                dest_cell = ws.cell(row=r_idx, column=col_idx)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.alignment = copy(src_cell.alignment)
                    dest_cell.number_format = src_cell.number_format
                    dest_cell.protection = copy(src_cell.protection)

            # Merge cells for countermeasure row structure
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=3, end_column=13)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=14, end_column=19)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=20, end_column=23)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=24, end_column=28)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=29, end_column=32)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=33, end_column=37)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=38, end_column=41)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=42, end_column=47)
            ws.merge_cells(start_row=r_idx, end_row=r_idx, start_column=52, end_column=55)

    # Populate Countermeasures
    for idx, p in enumerate(countermeasures):
        r_idx = start_row + idx
        set_cell(f"C{r_idx}", p["what"] or "")
        set_cell(f"N{r_idx}", p["why"] or "")
        set_cell(f"T{r_idx}", p["how"] or "")
        set_cell(f"X{r_idx}", p["where_place"] or "")

        plan_when = p["when_date"]
        if plan_when:
            if isinstance(plan_when, (datetime, date)):
                set_cell(f"AC{r_idx}", plan_when.strftime("%d/%m/%Y"))
            else:
                set_cell(f"AC{r_idx}", str(plan_when))

        set_cell(f"AG{r_idx}", p["who"] or "")

        if p["id_control_hierarchy"]:
            hier = next((h for h in hierarchies if h["id"] == p["id_control_hierarchy"]), None)
            set_cell(f"AL{r_idx}", hier["abbreviation"] if hier and "abbreviation" in hier else (hier["name"] if hier else ""))

        if p["id_verification_method"]:
            meth = next((m for m in methods if m["id"] == p["id_verification_method"]), None)
            set_cell(f"AP{r_idx}", meth["name"] if meth else "")

        set_cell(f"AV{r_idx}", "X" if p["ok"] else "" )
        set_cell(f"AX{r_idx}", "X" if p["ng"] else "" )
        set_cell(f"AZ{r_idx}", p["comment"] or p.get("comments") or "")

    # Participants (desplazamiento por extra_rows de countermeasures + factor_shift)
    shift_amount = extra_rows + factor_shift
    workers = [p for p in participants if p["participant_type"] == 'Participante']
    interested = [p for p in participants if p["participant_type"] == 'Parte interesada pertinente']
    reps = [p for p in participants if p["participant_type"] == 'Representante de los trabajadores']

    # Worker Participants
    r_worker = 101 + shift_amount
    for p in workers:
        set_cell(f"C{r_worker}", p["name"] or "")
        set_cell(f"L{r_worker}", p["department"] or "")
        if p["id_cost_center"]:
            cc = next((c for c in cost_centers if c["id"] == p["id_cost_center"]), None)
            set_cell(f"S{r_worker}", cc["name"] if cc else "")
        r_worker += 1

    # Interested Parties
    r_interested = 101 + shift_amount
    for p in interested:
        set_cell(f"V{r_interested}", p["name"] or "")
        set_cell(f"AD{r_interested}", p["department"] or "")
        if p["id_cost_center"]:
            cc = next((c for c in cost_centers if c["id"] == p["id_cost_center"]), None)
            set_cell(f"AI{r_interested}", cc["name"] if cc else "")
        r_interested += 1

    # Representatives
    r_rep = 101 + shift_amount
    for p in reps:
        set_cell(f"AN{r_rep}", p["name"] or "")
        set_cell(f"AV{r_rep}", p["department"] or "")
        if p["id_cost_center"]:
            cc = next((c for c in cost_centers if c["id"] == p["id_cost_center"]), None)
            set_cell(f"BA{r_rep}", cc["name"] if cc else "")
        r_rep += 1

    # Save to BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    # Modify drawings in zip bytes
    zip_bytes = out.getvalue()
    modified_zip_bytes = modify_excel_zip_with_drawings_py(zip_bytes, incident["injury"] or "", factors, template_path)
    out_modified = BytesIO(modified_zip_bytes)

    filename = f"reporte_{incident['incident_folio'] or id_incident}.xlsx"
    return StreamingResponse(
        out_modified,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def modify_excel_zip_with_drawings_py(zip_bytes, injury_text, factors, template_path):
    import zipfile
    import io

    # Extract modified sheets/styles from the openpyxl zip
    in_file = io.BytesIO(zip_bytes)
    try:
        with zipfile.ZipFile(in_file, 'r') as z_openpyxl:
            sheet1_data = z_openpyxl.read('xl/worksheets/sheet1.xml')
            styles_data = z_openpyxl.read('xl/styles.xml')
    except Exception as e:
        print("Error reading from openpyxl zip:", e)
        return zip_bytes

    # Rebuild using the original template zip as base to preserve all drawing files, media, and relationship files
    out_file = io.BytesIO()
    try:
        with zipfile.ZipFile(template_path, 'r') as z_orig:
            with zipfile.ZipFile(out_file, 'w', compression=zipfile.ZIP_DEFLATED) as z_out:
                for item in z_orig.infolist():
                    if item.filename == 'xl/worksheets/sheet1.xml':
                        z_out.writestr(item.filename, sheet1_data)
                    elif item.filename == 'xl/styles.xml':
                        z_out.writestr(item.filename, styles_data)
                    elif item.filename == 'xl/drawings/drawing1.xml':
                        xml_str = z_orig.read(item.filename).decode('utf-8')
                        modified_xml = modify_drawing_xml_py(xml_str, injury_text, factors)
                        z_out.writestr(item.filename, modified_xml.encode('utf-8'))
                    else:
                        z_out.writestr(item.filename, z_orig.read(item.filename))
        return out_file.getvalue()
    except Exception as e:
        print("Error rebuilding zip with original template base in Python:", e)
        return zip_bytes

def modify_drawing_xml_py(xml_str, injury_text, factors):
    import re
    # Extract all twoCellAnchor tags
    anchors = re.findall(r'<xdr:twoCellAnchor[\s\S]*?</xdr:twoCellAnchor>', xml_str)
    
    if len(anchors) < 3:
        return xml_str
        
    # We only keep the top header logos/texts (anchors 0, 1, 2)
    # The factor tree shapes/diagram are completely discarded from the drawing layer
    new_anchors = anchors[:3]
    
    header_idx = xml_str.find("<xdr:twoCellAnchor")
    xml_header = xml_str[:header_idx] if header_idx != -1 else ""
    footer_idx = xml_str.rfind("</xdr:twoCellAnchor>")
    xml_footer = xml_str[footer_idx + len("</xdr:twoCellAnchor>"):] if footer_idx != -1 else "</xdr:wsDr>"
    
    return xml_header + "\n".join(new_anchors) + xml_footer

def shift_anchor_row_py(anchor_str, offset):
    import re
    if offset == 0:
        return anchor_str
    def replace_row(match):
        val = int(match.group(1))
        return f"<xdr:row>{val + offset}</xdr:row>"
    return re.sub(r'<xdr:row>(\d+)</xdr:row>', replace_row, anchor_str)

def set_shape_id_py(anchor_str, new_id):
    import re
    return re.sub(r'<xdr:cNvPr id="(\d+)"', f'<xdr:cNvPr id="{new_id}"', anchor_str)

def set_shape_text_py(anchor_str, text):
    import re
    if "<a:t>" in anchor_str:
        return re.sub(r'<a:t>[\s\S]*?</a:t>', f'<a:t>{text}</a:t>', anchor_str)
    elif text:
        new_run = f'<a:r><a:rPr lang="es-MX" sz="1000"/><a:t>{text}</a:t></a:r>'
        return re.sub(r'<a:p>([\s\S]*?)</a:p>', f'<a:p>{new_run}</a:p>', anchor_str)
    return anchor_str

def make_judgment_shape_py(template_shapes, is_ok, offset, col_start, next_id):
    import re
    tpl_idx = 17 if is_ok else 18
    shape = template_shapes[tpl_idx]
    shape = set_shape_id_py(shape, next_id)
    shape = shift_anchor_row_py(shape, offset)
    
    state = {"count": 0, "from_col_val": 0}
    def replace_col(match):
        state["count"] += 1
        col_val = int(match.group(1))
        if state["count"] == 1:
            state["from_col_val"] = col_val
            return f"<xdr:col>{col_start}</xdr:col>"
        elif state["count"] == 2:
            span = col_val - state["from_col_val"]
            return f"<xdr:col>{col_start + span}</xdr:col>"
        return match.group(0)
        
    return re.sub(r'<xdr:col>(\d+)</xdr:col>', replace_col, shape)
